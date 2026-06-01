#!/usr/bin/env python3
"""Import Clients.xlsx + Clara's Clients.xlsx into dashboard JSON."""
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl

CLIENTS_XLSX = Path("/Users/oliver/Downloads/Clients.xlsx")
CLARA_XLSX = Path("/Users/oliver/Downloads/Clara's Clients.xlsx")
OUT = Path(__file__).parent / "data" / "excel-import.json"


def norm_name(s):
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def to_date(val):
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if val is None:
        return ""
    s = str(val).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    return s


def parse_sessions(pkg):
    if not pkg:
        return 0
    s = str(pkg)
    m = re.search(r"(\d+)\s*session", s, re.I)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d+)", s)
    return int(m.group(1)) if m else 0


def parse_amount(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("$", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def service_type(svc):
    s = (svc or "").strip().lower()
    if "lagree" in s:
        return "lagree"
    return "pilates"


def read_clients_kpis(ws):
    kpis = {}
    labels = {
        "revenue": "Revenue",
        "purchases": "Number of purchases",
        "aov": "Average Purchase Value",
        "uniqueCustomers": "Unique Customer",
        "purchaseFrequency": "Purchase Frequency",
        "churnRate": "Churn Rate",
        "customersLost": "Customers Lost",
        "customerLifespan": "Customer Lifespan",
        "ltv": "LTV",
    }
    rev_map = {
        "Revenue": "revenue",
        "Number of purchases": "purchases",
        "Average Purchase Value ( AOV )": "aov",
        "Average Purchase Value": "aov",
        "Unique Customer": "uniqueCustomers",
        "Purchase Frequency": "purchaseFrequency",
        "Churn Rate": "churnRate",
        "Customers Lost": "customersLost",
        "Customer Lifespan": "customerLifespan",
        "LTV": "ltv",
    }
    for r in range(1, 15):
        label = ws.cell(r, 12).value
        val = ws.cell(r, 13).value
        if not label or val is None:
            continue
        label = str(label).strip()
        key = rev_map.get(label)
        if not key and label in labels.values():
            key = [k for k, v in labels.items() if v == label][0]
        if key:
            kpis[key] = float(val) if isinstance(val, (int, float)) else val
    return kpis


def import_client_packages():
    wb = openpyxl.load_workbook(CLIENTS_XLSX, data_only=True)
    ws = wb["Client Packages"]
    kpis = read_clients_kpis(ws)
    payments = []
    clients = {}
    pid = 1

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 4).value
        amount = ws.cell(r, 8).value
        if not name or not isinstance(amount, (int, float)):
            continue
        name = str(name).strip()
        amt = float(amount)
        if amt <= 0:
            continue
        dt = to_date(ws.cell(r, 2).value)
        if not dt:
            continue
        month = str(ws.cell(r, 1).value or "")[:20]
        svc = ws.cell(r, 3).value
        package = str(ws.cell(r, 5).value or "").strip()
        date_from = to_date(ws.cell(r, 6).value)
        date_to = to_date(ws.cell(r, 7).value)
        sessions = parse_sessions(package)

        pay = {
            "id": pid,
            "name": name,
            "amount": round(amt, 2),
            "paid": True,
            "date": dt,
            "plan": package or "Package",
            "notes": f"{month} · {svc or 'Pilates'}".strip(" · "),
            "service": svc or "Pilates",
            "from": date_from,
            "to": date_to,
            "sessions": sessions,
            "source": "clients",
            "at": f"{dt}T12:00:00.000Z",
        }
        payments.append(pay)
        pid += 1

        nk = norm_name(name)
        if nk not in clients:
            clients[nk] = {
                "name": name,
                "totalRevenue": 0,
                "purchaseCount": 0,
                "packages": [],
                "type": service_type(svc),
            }
        c = clients[nk]
        c["totalRevenue"] = round(c["totalRevenue"] + amt, 2)
        c["purchaseCount"] += 1
        c["packages"].append(
            {
                "date": dt,
                "package": package,
                "amount": round(amt, 2),
                "from": date_from,
                "to": date_to,
                "sessions": sessions,
            }
        )
        c["lastPackage"] = package
        c["lastAmount"] = round(amt, 2)
        c["lastDate"] = dt
        c["membershipFee"] = round(amt, 2)

    return kpis, payments, clients


def clara_row_amount(package, c_cut, s_cut):
    sessions = parse_sessions(package)
    if sessions <= 0:
        sessions = 1
    cc = parse_amount(c_cut)
    sc = parse_amount(s_cut)
    if cc or sc:
        return round(sessions * (cc + sc), 2)
    return 0.0


def import_clara_clients():
    wb = openpyxl.load_workbook(CLARA_XLSX, data_only=True)
    ws = wb["Claras Client"]
    payments = []
    pid = 100000

    for r in range(4, ws.max_row + 1):
        name = ws.cell(r, 3).value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        package = str(ws.cell(r, 4).value or "").strip()
        dt = to_date(ws.cell(r, 2).value)
        if not dt:
            continue
        month = str(ws.cell(r, 1).value or "")
        c_cut = ws.cell(r, 7).value
        s_cut = ws.cell(r, 8).value
        amt = clara_row_amount(package, c_cut, s_cut)
        if amt <= 0:
            continue
        date_from = to_date(ws.cell(r, 5).value)
        date_to = to_date(ws.cell(r, 6).value)
        sessions = parse_sessions(package) or 1

        payments.append(
            {
                "id": pid,
                "name": name,
                "amount": amt,
                "paid": True,
                "date": dt,
                "plan": f"Clara: {package}" if package else "Clara session",
                "notes": f"{month} · 50% Clara ${parse_amount(c_cut)} · 50% Studio ${parse_amount(s_cut)}",
                "from": date_from,
                "to": date_to,
                "sessions": sessions,
                "claraCut": parse_amount(c_cut) * sessions,
                "studioCut": parse_amount(s_cut) * sessions,
                "source": "clara",
                "at": f"{dt}T12:00:00.000Z",
            }
        )
        pid += 1

    return payments


def build_ledger(payments):
    ledger = {}
    for p in payments:
        mk = p["date"][:7]
        ledger.setdefault(mk, []).append(
            {
                "id": p["id"],
                "name": p["name"],
                "amount": p["amount"],
                "paid": p["paid"],
                "notes": p.get("notes", ""),
                "at": p["at"],
            }
        )
    return ledger


def main():
    kpis, main_payments, clients = import_client_packages()
    clara_payments = import_clara_clients()

    # Dedupe clara vs main: same name+date+similar amount
    main_keys = {
        f"{norm_name(p['name'])}|{p['date']}|{round(p['amount'])}"
        for p in main_payments
    }
    clara_unique = []
    for p in clara_payments:
        k = f"{norm_name(p['name'])}|{p['date']}|{round(p['amount'])}"
        if k in main_keys:
            p["notes"] = (p.get("notes") or "") + " · Clara instructor"
            continue
        clara_unique.append(p)

    all_payments = main_payments + clara_unique
    all_payments.sort(key=lambda x: (x["date"], x["name"]))

    # Client session totals from packages
    for nk, c in clients.items():
        total_sess = sum(p["sessions"] for p in c["packages"])
        c["totalSessions"] = total_sess
        c["completedSessions"] = total_sess  # purchased capacity
        c["membership"] = c.get("lastPackage") or "Member"

    clara_revenue = sum(p["amount"] for p in clara_payments)
    clara_cut_total = sum(p.get("claraCut", 0) for p in clara_payments)

    out = {
        "version": 1,
        "importedAt": datetime.utcnow().isoformat() + "Z",
        "kpis": kpis,
        "stats": {
            "paymentCount": len(all_payments),
            "clientCount": len(clients),
            "claraPaymentCount": len(clara_payments),
            "claraRevenue": round(clara_revenue, 2),
            "claraCutTotal": round(clara_cut_total, 2),
            "totalRevenue": round(sum(p["amount"] for p in all_payments), 2),
        },
        "payments": all_payments,
        "membershipLedger": build_ledger(all_payments),
        "clients": clients,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False)
    print(f"Wrote {OUT}")
    print(f"  Payments: {len(all_payments)} ({len(main_payments)} studio + {len(clara_unique)} Clara-only)")
    print(f"  Clients: {len(clients)}")
    print(f"  KPI revenue: {kpis.get('revenue')}")


if __name__ == "__main__":
    main()
